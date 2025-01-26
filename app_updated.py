
import streamlit as st
from sentence_transformers import SentenceTransformer, util
import PyPDF2
from docx import Document
import openpyxl
import os
import pandas as pd

# Initialize the model
@st.cache_resource
def load_model():
    return SentenceTransformer('all-MiniLM-L6-v2')

model = load_model()

# File handling functions
def extract_text_from_pdf(file):
    try:
        pdf_reader = PyPDF2.PdfReader(file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text()
        return text
    except Exception as e:
        return f"Error reading PDF file: {str(e)}"

def extract_text_from_docx(file):
    try:
        doc = Document(file)
        return "\n".join([paragraph.text for paragraph in doc.paragraphs])
    except Exception as e:
        return f"Error reading Word document: {str(e)}"

def extract_text_from_excel(file):
    try:
        workbook = openpyxl.load_workbook(file, data_only=True)
        text = ""
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                text += " ".join([str(cell) if cell else "" for cell in row]) + "\n"
        return text
    except Exception as e:
        return f"Error reading Excel file: {str(e)}"

def process_file(file, file_name):
    if file_name.endswith(".pdf"):
        return extract_text_from_pdf(file)
    elif file_name.endswith(".docx"):
        return extract_text_from_docx(file)
    elif file_name.endswith(".xlsx"):
        return extract_text_from_excel(file)
    elif file_name.endswith(".txt"):
        return str(file.read(), "utf-8")
    else:
        return None

# File upload and indexing
st.title("AI Document Search Agent")

uploaded_files = []
uploaded_folder = st.file_uploader("Upload files or folders (zip format for folders)", accept_multiple_files=True)
if uploaded_folder:
    st.success(f"{len(uploaded_folder)} file(s) uploaded successfully!")
    for file in uploaded_folder:
        uploaded_files.append((file.name, process_file(file, file.name)))

# File Management Dashboard
st.header("File Management")
if uploaded_files:
    st.write("Uploaded Files:")
    for file_name, file_content in uploaded_files:
        st.markdown(f"- **{file_name}**")

# Query and Search
st.header("Query and Search")
query = st.text_input("Enter your search query:")
if query:
    document_texts = [content for _, content in uploaded_files if content]
    
    # Validate input before generating embeddings
    if len(document_texts) == 0:
        st.warning("No documents are available for search. Please upload files first.")
    elif not query.strip():
        st.warning("Please enter a valid search query.")
    else:
        # Generate embeddings
        embeddings = model.encode(document_texts, convert_to_tensor=True)
        query_embedding = model.encode(query, convert_to_tensor=True)
        
        # Perform similarity search with validation
        if embeddings.size(0) == 0 or query_embedding.size(0) == 0:
            st.warning("Could not generate embeddings. Ensure the query and files are valid.")
        else:
            scores = util.pytorch_cos_sim(query_embedding, embeddings)[0]
            results = sorted(zip(uploaded_files, scores), key=lambda x: x[1], reverse=True)

            # Display results
            st.subheader("Search Results:")
            for (file_name, file_content), score in results[:5]:  # Top 5 results
                st.markdown(f"### {file_name} (Score: {score.item():.4f})")
                st.write(file_content[:500] + "...")  # Preview the first 500 characters

# Query Logging
st.header("Query History")
query_log = []
if "query_log" not in st.session_state:
    st.session_state.query_log = []
if query:
    st.session_state.query_log.append({"Query": query, "Results": results[:5]})

if st.session_state.query_log:
    df = pd.DataFrame(st.session_state.query_log)
    st.write(df)
    if st.button("Export Query Log"):
        csv = df.to_csv(index=False).encode("utf-8")
        st.download_button(label="Download Query Log", data=csv, file_name="query_log.csv", mime="text/csv")
